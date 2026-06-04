import os
from openpyxl import load_workbook
from datetime import datetime, date

ARCHIVO_EXCEL = os.path.join(
    os.path.dirname(__file__),
    "BBDD_empleados.xlsx"
)


def cargar_excel():
    try:
        return load_workbook(ARCHIVO_EXCEL)

    except FileNotFoundError:
        print("Error: no se encontró el archivo Excel.")
        return None

    except PermissionError:
        print("Error: el archivo Excel está abierto.")
        return None

    except Exception as error:
        print(f"Error al abrir Excel: {error}")
        return None


def guardar_archivo(archivo):
    try:
        archivo.save(ARCHIVO_EXCEL)
        return True

    except PermissionError:
        print("\nError: cierre el archivo Excel antes de guardar.")
        return False

    except Exception as error:
        print(f"\nError al guardar Excel: {error}")
        return False


def buscar_empleado(hoja, legajo):
    for fila in range(2, hoja.max_row + 1):
        if str(hoja.cell(fila, 1).value) == str(legajo):
            return fila

    return None


def solicitar_fecha(mensaje):
    while True:
        fecha_texto = input(mensaje).strip()

        try:
            fecha = datetime.strptime(fecha_texto, "%d/%m/%Y")
            return fecha

        except ValueError:
            print("Error: formato inválido. Utilice DD/MM/AAAA.")


def solicitar_rango_fechas():
    while True:
        fecha_inicio = solicitar_fecha(
            "\nIngrese fecha de inicio (DD/MM/AAAA): "
        )

        fecha_fin = solicitar_fecha(
            "Ingrese fecha de fin (DD/MM/AAAA): "
        )

        if fecha_inicio.date() < datetime.today().date():
            print("\nBot: No puede solicitar vacaciones con fecha anterior a hoy.")
            print("Bot: Ingrese nuevamente las fechas.")
            continue

        if fecha_fin < fecha_inicio:
            print("\nBot: La fecha de fin debe ser igual o posterior a la fecha de inicio.")
            print("Bot: Ingrese nuevamente las fechas.")
            continue

        return fecha_inicio, fecha_fin


def solicitar_decision_supervisor():
    while True:
        decision = input(
            "\nSupervisor, ¿aprueba la solicitud? (S/N): "
        ).strip().upper()

        if decision in ["S", "N"]:
            return decision

        print("Error: debe ingresar S o N.")


def solicitar_motivo_rechazo():
    while True:
        motivo = input(
            "\nSupervisor, ingrese el motivo del rechazo: "
        ).strip()

        if motivo != "":
            return motivo

        print("Error: debe ingresar un motivo de rechazo.")


def solicitar_confirmacion_usuario():
    while True:
        respuesta = input(
            "\nEmpleado, ¿desea continuar con la solicitud? (S/N): "
        ).strip().upper()

        if respuesta in ["S", "N"]:
            return respuesta

        print("Error: debe ingresar S o N.")


def supervisor_esta_de_vacaciones(fecha_inicio, fecha_fin):
    anio = fecha_inicio.year

    inicio_vacaciones_supervisor = date(anio, 9, 15)
    fin_vacaciones_supervisor = date(anio, 10, 20)

    inicio_solicitud = fecha_inicio.date()
    fin_solicitud = fecha_fin.date()

    return (
        inicio_solicitud <= fin_vacaciones_supervisor
        and fin_solicitud >= inicio_vacaciones_supervisor
    )


def registrar_solicitud(
    hoja_historico,
    legajo,
    fecha_inicio,
    fecha_fin,
    dias_solicitados,
    estado,
    aprobado_por,
    comentario
):
    nuevo_id = hoja_historico.max_row

    hoja_historico.append([
        nuevo_id,
        int(legajo),
        datetime.now().strftime("%d/%m/%Y"),
        fecha_inicio.strftime("%d/%m/%Y"),
        fecha_fin.strftime("%d/%m/%Y"),
        dias_solicitados,
        estado,
        aprobado_por,
        comentario
    ])


def aprobar_por_sistema(
    archivo,
    hoja_empleados,
    hoja_historico,
    fila_empleado,
    legajo,
    fecha_inicio,
    fecha_fin,
    dias_solicitados,
    dias_disponibles
):
    dias_usados = hoja_empleados.cell(fila_empleado, 6).value
    nuevo_disponible = dias_disponibles - dias_solicitados

    hoja_empleados.cell(fila_empleado, 6).value = dias_usados + dias_solicitados
    hoja_empleados.cell(fila_empleado, 7).value = nuevo_disponible

    registrar_solicitud(
        hoja_historico,
        legajo,
        fecha_inicio,
        fecha_fin,
        dias_solicitados,
        "Aprobada",
        "Sistema",
        "Vacaciones aprobadas por sistema, porque el supervisor se encuentra de vacaciones"
    )

    if guardar_archivo(archivo):
        print("\nBot: Solicitud aprobada automáticamente.")
        print("Bot: Responsable: Sistema.")
        print("Bot: Motivo: vacaciones aprobadas por sistema, porque el supervisor se encuentra de vacaciones.")
        print(f"Bot: Nuevo saldo disponible: {nuevo_disponible} días.")


def rechazar_por_sistema_supervisor_vacaciones(
    archivo,
    hoja_historico,
    legajo,
    fecha_inicio,
    fecha_fin,
    dias_solicitados
):
    registrar_solicitud(
        hoja_historico,
        legajo,
        fecha_inicio,
        fecha_fin,
        dias_solicitados,
        "Rechazada",
        "Sistema",
        "Solicitud rechazada por falta de personal, porque el supervisor se encuentra de vacaciones"
    )

    if guardar_archivo(archivo):
        print("\nBot: Solicitud rechazada automáticamente.")
        print("Bot: Responsable: Sistema.")
        print("Bot: Motivo: falta de personal, porque el supervisor se encuentra de vacaciones.")


def chatbot_vacaciones():
    try:
        print("=" * 45)
        print(" BOT DE GESTIÓN DE VACACIONES ")
        print("=" * 45)

        archivo = cargar_excel()

        if archivo is None:
            return

        hoja_empleados = archivo["Empleados_vacaciones"]
        hoja_historico = archivo["Historico_vacaciones"]

        legajo = input("\nIngrese su legajo: ").strip()

        if legajo == "":
            print("\nBot: Debe ingresar un legajo.")
            return

        if not legajo.isdigit():
            print("\nBot: El legajo debe ser numérico.")
            return

        fila_empleado = buscar_empleado(hoja_empleados, legajo)

        if fila_empleado is None:
            print("\nBot: Legajo inexistente.")
            print("Bot: Solicitud rechazada.")
            return

        nombre = hoja_empleados.cell(fila_empleado, 2).value
        apellido = hoja_empleados.cell(fila_empleado, 3).value
        dias_disponibles = hoja_empleados.cell(fila_empleado, 7).value

        print(f"\nBot: Bienvenido/a {nombre} {apellido}.")
        print(f"Bot: Usted dispone de {dias_disponibles} días.")

        print("\n=== SOLICITUD DE VACACIONES ===")

        fecha_inicio, fecha_fin = solicitar_rango_fechas()

        dias_solicitados = (fecha_fin - fecha_inicio).days + 1

        print(f"\nBot: Días calculados automáticamente: {dias_solicitados}")
        print("\nBot: Validando información...")

        if dias_solicitados > dias_disponibles:
            registrar_solicitud(
                hoja_historico,
                legajo,
                fecha_inicio,
                fecha_fin,
                dias_solicitados,
                "Rechazada",
                "Sistema",
                "Saldo insuficiente"
            )

            guardar_archivo(archivo)

            print("\nBot: Solicitud rechazada automáticamente.")
            print("Bot: Responsable: Sistema.")
            print("Bot: Motivo: saldo insuficiente.")
            return

        if supervisor_esta_de_vacaciones(fecha_inicio, fecha_fin):
            print("\nBot: Atención.")
            print("Bot: El supervisor se encuentra de vacaciones entre el 15/09 y el 20/10.")
            print("Bot: En este período, el sistema evaluará automáticamente la solicitud.")

            continuar = solicitar_confirmacion_usuario()

            if continuar == "N":
                print("\nBot: Solicitud cancelada por el usuario.")
                return

            if dias_solicitados > 10:
                rechazar_por_sistema_supervisor_vacaciones(
                    archivo,
                    hoja_historico,
                    legajo,
                    fecha_inicio,
                    fecha_fin,
                    dias_solicitados
                )
                return

            aprobar_por_sistema(
                archivo,
                hoja_empleados,
                hoja_historico,
                fila_empleado,
                legajo,
                fecha_inicio,
                fecha_fin,
                dias_solicitados,
                dias_disponibles
            )
            return

        registrar_solicitud(
            hoja_historico,
            legajo,
            fecha_inicio,
            fecha_fin,
            dias_solicitados,
            "Pendiente",
            "",
            "Solicitud enviada al supervisor"
        )

        print("\nBot: Solicitud registrada con estado PENDIENTE.")
        print("Bot: Enviando solicitud al supervisor...")

        decision = solicitar_decision_supervisor()

        ultima_fila = hoja_historico.max_row

        if decision == "S":
            dias_usados = hoja_empleados.cell(fila_empleado, 6).value
            nuevo_disponible = dias_disponibles - dias_solicitados

            hoja_empleados.cell(fila_empleado, 6).value = dias_usados + dias_solicitados
            hoja_empleados.cell(fila_empleado, 7).value = nuevo_disponible

            hoja_historico.cell(ultima_fila, 7).value = "Aprobada"
            hoja_historico.cell(ultima_fila, 8).value = "Supervisor"
            hoja_historico.cell(ultima_fila, 9).value = "Solicitud aprobada"

            if guardar_archivo(archivo):
                print("\nBot: Solicitud aprobada.")
                print(f"Bot: Nuevo saldo disponible: {nuevo_disponible} días.")

        else:
            motivo_rechazo = solicitar_motivo_rechazo()

            hoja_historico.cell(ultima_fila, 7).value = "Rechazada"
            hoja_historico.cell(ultima_fila, 8).value = "Supervisor"
            hoja_historico.cell(ultima_fila, 9).value = motivo_rechazo

            if guardar_archivo(archivo):
                print("\nBot: Solicitud rechazada por el supervisor.")
                print(f"Bot: Motivo del rechazo: {motivo_rechazo}")

    except KeyError as error:
        print(f"\nError: no existe la hoja indicada en el Excel: {error}")

    except Exception as error:
        print(f"\nError inesperado del sistema: {error}")


chatbot_vacaciones()